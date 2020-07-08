#爬取指定达人的相关数据并存储
import requests
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook


#获取粉丝量、标签、转赞评数据、
def get_info(html):#html为页面源代码
    try:
        #获取标签及粉丝数
        bs = BeautifulSoup(source, 'lxml')
        biaoqian=bs.find('p',class_='mod-fil-desc m-text-cut').get_text().strip()
        fensishu=bs.find('div', class_='mod-fil-fans').find_all('span')[1].string
        print('标签：'+biaoqian+'\n'+'粉丝数：'+fensishu+'\n')
    except:
        print('标签和粉丝部分有问题')
    try:
        info=bs.find_all('footer',class_='m-ctrl-box m-box-center-a')

        #print('转发' + '\t评论' + '\t点赞')
        zf_list=[]
        pl_list=[]
        dz_list=[]
        for x in range(30):
            zf = info[x].find_all('div',class_='m-diy-btn m-box-col m-box-center m-box-center-a')[0].get_text().strip()
            pl = info[x].find_all('div',class_='m-diy-btn m-box-col m-box-center m-box-center-a')[1].get_text().strip()
            dz = info[x].find_all('div',class_='m-diy-btn m-box-col m-box-center m-box-center-a')[2].get_text().strip()
            zf_list.append(zf)
            pl_list.append(pl)
            dz_list.append(dz)
            #print(zf+'\t'+pl+'\t'+dz+'\n')
        print('转发'+zf_list+'\n'+'评论'+pl_list+'\n'+'点赞'+dz_list)
    except:
        print('数据获取部分有问题')
    return zf_list,pl_list,dz_list


#微博的平均转评赞
def get_avg(zpz):
    try:
        num = len(zpz[0])#获取转评赞数量
        print('最近'+str(num)+'条微博的平均数据如下'+'\n')
        all = 0
        '''avg_zf = 0
        avg_pl = 0
        avg_dz = 0'''
        for a in range(0,3):
            if a == 0:
                for x in range(0,num):
                    if zpz[a][x] =='转发':
                        zpz[a][x] = float('0')
                    elif ('万' in str(zpz[a][x])):
                        zpz[a][x] = float(zpz[a][x].replace('万',''))*10000
                    else:
                        zpz[a][x] = float(zpz[a][x])
                    all = all + zpz[a][x]
                   # avg_zf = all
                print('平均转发'+str(all/num))
                all=0
            elif a==1:
                for x in range(0,num):
                    if zpz[a][x] =='评论':
                        zpz[a][x] = float('0')
                    elif ('万' in str(zpz[a][x])):
                        zpz[a][x] = float(zpz[a][x].replace('万',''))*10000
                    else:
                        zpz[a][x] = float(zpz[a][x])
                    all = all + zpz[a][x]
                   # avg_pl=all
                print('平均评论' + str(all / num))
                all=0
            else:
                for x in range(0,num):
                    if zpz[a][x] =='赞':
                        zpz[a][x] = float('0')
                    elif ('万' in str(zpz[a][x])):
                        zpz[a][x] = float(zpz[a][x].replace('万',''))*10000
                    else:
                        zpz[a][x] = float(zpz[a][x])
                    all = all + zpz[a][x]
                    #avg_dz = all
                print('平均点赞' + str(all / num))
    except:
        print("get_avg有问题")

#写入Excel
def save_excel(name,kol_url,zpz):
    wb = Workbook()
    ws = wb.active
    ws['a1']='微博昵称'
    ws['b1']='微博标签'
    ws['c1'] = '微博链接'
    ws['d1']='近30条平均转发数'
    ws['e1']='近30条平均评论数'
    ws['f1']='近30条平均点赞数'
    ws.append([])
    '''
    #输出近30条微博的转评赞数据到Excel
    for c in range(0,3):
        for r in range(0,30):
            ws.cell(row=r+2,column=c+1,value=zpz[c][r])
    '''
    wb.save('/Users/wkc/Desktop/微博数据.xlsx')

#下滑页面
def scroll_to_bottom(driver):
    js = "return action=document.body.scrollHeight"
    # 初始化现在滚动条所在高度为0
    height = 0
    # 当前窗口总高度
    new_height = driver.execute_script(js)
    first_time =time.time()
    while height < new_height:
        # 将滚动条调整至页面底部
        for i in range(height, new_height, 150):
            driver.execute_script('window.scrollTo(0, {})'.format(i))
            time.sleep(0.5)
        height = new_height
        time.sleep(2)
        new_height = driver.execute_script(js)
        if time.time()-first_time>20:
            break

#打开指定网页获取页面代码和kol主页链接
def get_html(url,name):#url为链接
    #无界面启动
    opt = Options()
    opt.headless = True
    driver = webdriver.Chrome(options=opt)

    #driver = webdriver.Safari()

    driver.get(url)
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[1]/div/div/div[2]/form/input').send_keys(name)
    driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[1]/div/div/div[2]/form/input').send_keys(Keys.ENTER)
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/div[1]/div/div/div/ul/li[2]/span').click()
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[4]/div/div/div[1]/div/div/div/div[1]').click()
    time.sleep(3)
    #滑动页面
    scroll_to_bottom(driver)
    time.sleep(3)
    source = driver.page_source
    #print(source)
    kol_url = driver.current_url

    return source,kol_url


if __name__ == "__main__":
    try:
        name=input('输入查找的kol：')
        time.sleep(2)
        url='https://m.weibo.cn/search?containerid=231583'

        g_h=get_html(url,name)
        source=g_h[0]
        kol_url=g_h[1]
        print(kol_url)
        a=get_info(source)
        get_avg(a)
       # save_excel(a)
    except:
        print("有错误！")
