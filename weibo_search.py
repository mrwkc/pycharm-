import urllib3
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

text=input("输入搜索内容：")
headers={
    'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'
    'Cookie: SINAGLOBAL=5671536165295.199.1576468457320; UOR=,,login.sina.com.cn; _s_tentry=login.sina.com.cn; Apache=7826497085699.521.1593999590505; ULV=1593999590519:9:2:1:7826497085699.521.1593999590505:1593739314638; login_sid_t=b5b8d03cf96a46915054eaea6866e619; cross_origin_proto=SSL; SCF=AnPtvZqpp2CTe0EZN4Kx4HSoTW9K-0LErdJKafODj7-rRN7K7uh56k9h7ibSCdqP1NNuFAxa6c9iNHU8MrNpeYk.; SUB=_2A25yBvE1DeRhGeBG7FIW-SnEwz-IHXVRcmX9rDV8PUNbmtAKLVSgkW9NQekj6l98lr4p_RMuNFg7pFGa7JD0UvgH; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5s-JhbyUNj2_0URbNskXe65JpX5KzhUgL.FoqRS05N1KMR1he2dJLoI05pehnLxK-L12qLBoMLxKqLBo5LBoBLxK-L1hqL1-zLxKqL12qL1K5LxKML1-BLBKet; SUHB=0wnFB544UwTlr4; ALF=1625535716; SSOLoginState=1593999717; wvr=6; webim_unReadCount=%7B%22time%22%3A1594001058815%2C%22dm_pub_total%22%3A0%2C%22chat_group_client%22%3A0%2C%22chat_group_notice%22%3A0%2C%22allcountNum%22%3A0%2C%22msgbox%22%3A0%7D; WBStorage=42212210b087ca50|undefined'
}

wb = Workbook()
ws = wb.active
count=0
for x in range(1):
    #url='https://s.weibo.com/user?q='+text+'&Refer=weibo_user&page='+str(x)   #找人
    url='https://s.weibo.com/user/&tag='+text+'&page='+str(x)                      #标签
    r = requests.get(url,headers=headers)
    bs =BeautifulSoup(r.text,'lxml')
    weibo_info=bs.find_all('div',class_='card card-user-b s-pg16 s-brt1')
    num=len(weibo_info) #一页的账号数量
    for i in range(0,num):
        print("第"+str(count+i+1)+'个')
        
        print(weibo_info[i].find_all('div')[1].find('a').get_text())   #输出微博名称
        ws.cell(row=count+i+1, column=1, value=weibo_info[i].find_all('div')[1].find('a').get_text())
        #print('https:'+weibo_info[i].find('div',class_='info').find('a')['href'])  #输出微博链接
        try:
            p_tag= weibo_info[i].find_all('div')[1].find_all('p')
            print(len(p_tag))
            print(p_tag[0].get_text().strip())
            for y in len(p_tag):

                if y == 0:
                    print(weibo_info[i].find_all('div')[1].find_all('p')[y].find('i').get_text())
                    ws.cell(row=count+i+1, column=2, value=weibo_info[i].find_all('div')[1].find_all('p')[y].find('i').get_text())
                    print(weibo_info[i].find_all('div')[1].find_all('p')[y].find('a').get_text()
                          +':'+'https:'+weibo_info[i].find_all('div')[1].find_all('p')[y].find('a')['href'])
                    ws.cell(row=count+i+1, column=3, value='https:'+weibo_info[i].find_all('div')[1].find_all('p')[y].find('a')['href'])
                elif y == 2:
                    print(weibo_info[i].find_all('div')[1].find_all('p')[y].get_text().strip())
                    ws.cell(row=count+i+1,column=4,value=weibo_info[i].find_all('div')[1].find_all('p')[y].find_all('span')[0].get_text())
                    ws.cell(row=count+i+1,column=5,value=weibo_info[i].find_all('div')[1].find_all('p')[y].find_all('span')[1].get_text())
                    ws.cell(row=count+i+1,column=6,value=weibo_info[i].find_all('div')[1].find_all('p')[y].find_all('span')[2].get_text())

                else:
                    print(weibo_info[i].find_all('div')[1].find_all('p')[y].get_text().strip())         #关注、粉丝、微博量
                    ws.cell(row=count+i+1, column=7, value=weibo_info[i].find_all('div')[1].find_all('p')[y].get_text().strip())

            print('===========================================')
        except:
            print('有点小问题')
    count +=20

wb.save('/Users/wkc/Desktop/微博美食.xlsx')